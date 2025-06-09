import os
os.environ["TRANSFORMERS_NO_ADVISORY_WARNINGS"] = "1"
os.environ["TRANSFROMERS_VERBOSITY"] = "error"

import shutil
import cv2
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import torch
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
import unicodedata
from difflib import get_close_matches
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import argparse

CHOSUNG_LIST = ['ㄱ','ㄲ','ㄴ','ㄷ','ㄸ','ㄹ','ㅁ','ㅂ','ㅃ','ㅅ','ㅆ','ㅇ','ㅈ','ㅉ','ㅊ','ㅋ','ㅌ','ㅍ','ㅎ']
JUNGSUNG_LIST = ['ㅏ','ㅐ','ㅑ','ㅒ','ㅓ','ㅔ','ㅕ','ㅖ','ㅗ','ㅘ','ㅙ','ㅚ','ㅛ','ㅜ','ㅝ','ㅞ','ㅟ','ㅠ','ㅡ','ㅢ','ㅣ']
JONGSUNG_LIST = ['', 'ㄱ','ㄲ','ㄳ','ㄴ','ㄵ','ㄶ','ㄷ','ㄹ','ㄺ','ㄻ','ㄼ','ㄽ','ㄾ','ㄿ','ㅀ','ㅁ','ㅂ','ㅄ','ㅅ','ㅆ','ㅇ','ㅈ','ㅊ','ㅋ','ㅌ','ㅍ','ㅎ']

i = 1

def copy_excel_to_result(src_path, dst_folder):
    if not os.path.exists(dst_folder):
        os.makedirs(dst_folder)
    filename = os.path.basename(src_path)
    dst_path = os.path.join(dst_folder, filename)
    if os.path.exists(dst_path):
        os.remove(dst_path)
        print(f"Removed existing file: {dst_path}")
    shutil.copy(src_path, dst_path)
    print(f"Copied {src_path} to {dst_path}")

def write_names_to_excel(excel_path, results):
    print(f"Opening workbook: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    processed_results = []
    for result in results:
        parts = result.split()
        if len(parts) >= 2:
            processed_results.append(parts[0])
            processed_results.append(parts[1])
        else:
            processed_results.append(result)
            processed_results.append(result)

    seats = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'seat':
                seats.append(cell)

    for i, cell in enumerate(seats):
        if i >= len(processed_results):
            break
        x, y = cell.row, cell.column
        ws.cell(row=x, column=y, value=processed_results[i])
    wb.save(excel_path)
    print(f"Workbook saved: {excel_path}")

def split_hangul_char(ch):
    if not (0xAC00 <= ord(ch) <= 0xD7A3):
        return (ch, '', '')
    base = ord(ch) - 0xAC00
    cho = base // 588
    jung = (base % 588) // 28
    jong = base % 28
    return (CHOSUNG_LIST[cho], JUNGSUNG_LIST[jung], JONGSUNG_LIST[jong])

def split_hangul_string(string):
    return [split_hangul_char(ch) for ch in string]

def weighted_similarity(a_parts, b_parts):
    score = 0
    total = 0
    weights = (2.0, 1.0, 0.5)
    for a, b in zip(a_parts, b_parts):
        for i in range(3):
            total += weights[i]
            if a[i] == b[i]:
                score += weights[i]
    return score / total if total else 0

def compare_slices(a, b):
    if len(a) < len(b):
        a, b = b, a
    max_score = 0
    for i in range(len(a) - len(b) + 1):
        a_slice = a[i:i+len(b)]
        sim = weighted_similarity(split_hangul_string(a_slice), split_hangul_string(b))
        max_score = max(max_score, sim)
    return max_score

def hangul_similarity(a, b):
    if len(a) >= 5 or len(b) >= 5:
        return compare_slices(a, b)
    if len(a) != len(b):
        return 0
    return weighted_similarity(split_hangul_string(a), split_hangul_string(b))

def load_reference_names(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return [name.strip() for name in f.read().split() if name.strip()]

def correct_names(text, reference_names, debug=False):
    global i
    text = text.replace("(", "").replace(")", "").replace("위닥", "")
    words = text.split()
    corrected = []
    for word in words:
        best_match = max(reference_names, key=lambda ref: hangul_similarity(word, ref))
        sim = hangul_similarity(word, best_match)
        if sim > 0.7:
            corrected.append(f"{i}.{best_match}{{{word}}}" if debug else best_match)
        else:
            match = get_close_matches(word, reference_names, n=1, cutoff=0.4)
            if match:
                corrected.append(f"{i}.{match[0]}[{word}]" if debug else match[0])
            else:
                corrected.append(f"{i}.error[{word}]" if debug else "error")
        i += 1
    if len(words) == 1:
        i += 1
    return " ".join(corrected)

def detect_handwriting_regions(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    _, binary = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    binary = 255 - binary
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (30, 5))
    dilated = cv2.dilate(binary, kernel, iterations=1)
    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    boxes = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        if 1200 > w > 900 and 300 > h > 100:
            boxes.append((x, y, x + w, y + h))
    boxes = sorted(boxes, key=lambda box: (round(box[1] / 10), box[0]))
    return boxes

def recognize_and_visualize(image_path, name_file_path, output_path="./result", debug=False):
    processor = TrOCRProcessor.from_pretrained("ddobokki/ko-trocr")
    model = VisionEncoderDecoderModel.from_pretrained("ddobokki/ko-trocr")
    
    # GPU 사용 가능 여부 확인
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model.to(device)
    model.eval()
    
    reference_names = load_reference_names(name_file_path)
    image = cv2.imread(image_path)
    if image is None:
        print(f"이미지를 불러올 수 없습니다: {image_path}")
        return
    boxes = detect_handwriting_regions(image)
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    image_pil = Image.fromarray(image_rgb)
    draw = ImageDraw.Draw(image_pil) if debug else None
    font = ImageFont.truetype("./data/NanumGothic.ttf", 20) if debug else None
    results = []
    for (x1, y1, x2, y2) in boxes:
        roi = image[y1:y2, x1:x2]
        if roi.size == 0:
            continue
        roi_pil = Image.fromarray(cv2.cvtColor(roi, cv2.COLOR_BGR2RGB))
        pixel_values = processor(images=roi_pil, return_tensors="pt").pixel_values
        with torch.no_grad():
            generated_ids = model.generate(pixel_values, max_length=64)
        text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]
        text = unicodedata.normalize("NFC", text)
        corrected_text = correct_names(text, reference_names, debug=debug)
        results.append(corrected_text)
        if debug:
            draw.rectangle([x1, y1, x2, y2], outline="red", width=2)
            draw.text((x1, y1 - 25), corrected_text, fill="blue", font=font)
    write_names_to_excel(output_path+"/blank.xlsx", results)
    if debug:
        image_pil = image_pil.resize((image_pil.width // 2, image_pil.height // 2))
        image_pil.show()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run handwritten OCR and export to Excel")
    parser.add_argument("--image", required=True, help="Path to the input image")
    parser.add_argument("--output", default="./result", help="Directory to save the result Excel file")
    parser.add_argument("--debug", action="store_true", help="Enable image visualization for debugging")
    args = parser.parse_args()

    copy_excel_to_result("./data/blank.xlsx", args.output)
    recognize_and_visualize(args.image, "./data/names.txt", args.output, debug=args.debug)
